# ============================================
# FICHIER : export_excel.py
# RÔLE    : Génère un fichier Excel avec le format :
#           Lignes = Comptes, Colonnes = Mois (4 sous-colonnes)
# ============================================

import os
import sys
import pandas as pd
from datetime import datetime

current_file  = os.path.abspath(__file__)
models_dir    = os.path.dirname(current_file)
ia_server_dir = os.path.dirname(models_dir)
project_root  = os.path.dirname(ia_server_dir)
sys.path.insert(0, ia_server_dir)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

NAVY        = "0D1B2A"
TEAL        = "0891B2"
TEAL_LIGHT  = "E0F7FA"
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F5F7FA"
GRAY_MID    = "E2E8F0"
GREEN_BG    = "ECFDF5"
GREEN_TEXT  = "065F46"
RED_BG      = "FEF2F2"
RED_TEXT    = "991B1B"
ORANGE_BG   = "FFFBEB"
ORANGE_TEXT = "92400E"
HEADER_ROW  = "1A3A5C"

MOIS_FR = ['Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin',
           'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Decembre']

def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color=WHITE, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def generer_excel_predictions(classe: int, year_target: int, year_realise: int = None) -> bytes:
    """Génère un Excel avec les prédictions pour une classe donnée."""
    from models.consolidate_all import calculer_predictions_par_classe
    
    result = calculer_predictions_par_classe(
        classe=classe,
        year_target=year_target,
        year_realise=year_realise
    )
    
    if 'error' in result:
        raise ValueError(result['error'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Classe {classe}"
    
    classe_label = result.get('classe_label', f'Classe {classe}')
    
    _build_prediction_sheet(ws, result, classe_label)
    
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_prediction_sheet(ws, data, classe_label):
    """Construit la feuille avec le format Mois x 4 colonnes."""
    
    classe = data['classe']
    year_realise = data['annee_realise']
    year_prediction = data['annee_prediction']
    comptes = data['comptes']
    totaux_globaux = data['totaux_globaux']
    mois_list = data['mois']
    
    ws.merge_cells("A1:B1")
    ws["A1"] = f"TABLEAU DE PREDICTION BUDGETAIRE - {classe_label}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = _align("center")
    ws.row_dimensions[1].height = 32
    
    ws.merge_cells("A2:B2")
    ws["A2"] = f"Annee Realise: {year_realise} | Annee Prediction: {year_prediction} | Genere le: {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="94A3B8")
    ws["A2"].fill = _fill("0D2E4A")
    ws["A2"].alignment = _align("center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8
    
    start_row = 4
    
    ws.cell(row=start_row, column=1, value="Compte")
    ws.cell(row=start_row, column=2, value="Libelle")
    
    col = 3
    mois_header_row = start_row
    sub_header_row = start_row + 1
    
    for mois in mois_list:
        ws.cell(row=mois_header_row, column=col, value=mois)
        ws.merge_cells(start_row=mois_header_row, start_column=col, 
                      end_row=mois_header_row, end_column=col+3)
        
        ws.cell(row=sub_header_row, column=col, value=f"Realise {year_realise}")
        ws.cell(row=sub_header_row, column=col+1, value="Moy 5 ans")
        ws.cell(row=sub_header_row, column=col+2, value=f"Prediction {year_prediction}")
        ws.cell(row=sub_header_row, column=col+3, value="Variation %")
        
        for c in range(col, col+4):
            ws.cell(row=mois_header_row, column=c).font = _font(bold=True, size=10)
            ws.cell(row=mois_header_row, column=c).fill = _fill(TEAL)
            ws.cell(row=mois_header_row, column=c).alignment = _align("center")
            ws.cell(row=mois_header_row, column=c).border = _border(TEAL)
            
            ws.cell(row=sub_header_row, column=c).font = _font(bold=True, size=9)
            ws.cell(row=sub_header_row, column=c).fill = _fill(HEADER_ROW)
            ws.cell(row=sub_header_row, column=c).alignment = _align("center", wrap=True)
            ws.cell(row=sub_header_row, column=c).border = _border("0D2240")
        
        col += 4
    
    ws.cell(row=mois_header_row, column=col, value="TOTAL")
    ws.merge_cells(start_row=mois_header_row, start_column=col, 
                  end_row=mois_header_row, end_column=col+3)
    
    for c in range(col, col+4):
        ws.cell(row=mois_header_row, column=c).font = _font(bold=True, size=10)
        ws.cell(row=mois_header_row, column=c).fill = _fill(TEAL)
        ws.cell(row=mois_header_row, column=c).alignment = _align("center")
        ws.cell(row=mois_header_row, column=c).border = _border(TEAL)
    
    ws.cell(row=sub_header_row, column=col, value=f"Realise {year_realise}")
    ws.cell(row=sub_header_row, column=col+1, value="Moy 5 ans")
    ws.cell(row=sub_header_row, column=col+2, value=f"Prediction {year_prediction}")
    ws.cell(row=sub_header_row, column=col+3, value="Variation %")
    
    for c in range(col, col+4):
        ws.cell(row=sub_header_row, column=c).font = _font(bold=True, size=9)
        ws.cell(row=sub_header_row, column=c).fill = _fill(TEAL)
        ws.cell(row=sub_header_row, column=c).alignment = _align("center", wrap=True)
        ws.cell(row=sub_header_row, column=c).border = _border(TEAL)
    
    ws.row_dimensions[mois_header_row].height = 24
    ws.row_dimensions[sub_header_row].height = 30
    
    data_start_row = sub_header_row + 1
    
    for i, compte_data in enumerate(comptes):
        row = data_start_row + i
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        
        ws.cell(row=row, column=1, value=compte_data['account'])
        ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=9, color=TEAL)
        ws.cell(row=row, column=1).fill = _fill(TEAL_LIGHT)
        ws.cell(row=row, column=1).alignment = _align("center")
        ws.cell(row=row, column=1).border = _border(GRAY_MID)
        
        ws.cell(row=row, column=2, value=compte_data['libelle'])
        ws.cell(row=row, column=2).font = Font(name="Arial", size=9, color="1E293B")
        ws.cell(row=row, column=2).fill = _fill(bg)
        ws.cell(row=row, column=2).alignment = _align("left")
        ws.cell(row=row, column=2).border = _border(GRAY_MID)
        
        col = 3
        for mois_data in compte_data['donnees_mensuelles']:
            realise = mois_data['realise']
            moyenne = mois_data['moyenne']
            prediction = mois_data['prediction']
            variation = mois_data['variation']
            
            ws.cell(row=row, column=col, value=realise)
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = _fill(bg)
            ws.cell(row=row, column=col).alignment = _align("right")
            ws.cell(row=row, column=col).border = _border(GRAY_MID)
            
            ws.cell(row=row, column=col+1, value=moyenne)
            ws.cell(row=row, column=col+1).number_format = '#,##0'
            ws.cell(row=row, column=col+1).fill = _fill(bg)
            ws.cell(row=row, column=col+1).alignment = _align("right")
            ws.cell(row=row, column=col+1).border = _border(GRAY_MID)
            
            ws.cell(row=row, column=col+2, value=prediction)
            ws.cell(row=row, column=col+2).number_format = '#,##0'
            ws.cell(row=row, column=col+2).fill = _fill(TEAL_LIGHT)
            ws.cell(row=row, column=col+2).font = Font(name="Arial", bold=True, size=9, color=TEAL)
            ws.cell(row=row, column=col+2).alignment = _align("right")
            ws.cell(row=row, column=col+2).border = _border(TEAL)
            
            ws.cell(row=row, column=col+3, value=variation / 100)
            ws.cell(row=row, column=col+3).number_format = '+0.0%;-0.0%;0.0%'
            vbg = GREEN_BG if variation >= 0 else RED_BG
            vfg = GREEN_TEXT if variation >= 0 else RED_TEXT
            ws.cell(row=row, column=col+3).fill = _fill(vbg)
            ws.cell(row=row, column=col+3).font = Font(name="Arial", bold=True, size=9, color=vfg)
            ws.cell(row=row, column=col+3).alignment = _align("center")
            ws.cell(row=row, column=col+3).border = _border(GRAY_MID)
            
            col += 4
        
        t = compte_data['totaux']
        ws.cell(row=row, column=col, value=t['realise'])
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).font = Font(name="Arial", bold=True, size=9, color="1E293B")
        ws.cell(row=row, column=col).fill = _fill(GRAY_MID)
        ws.cell(row=row, column=col).alignment = _align("right")
        ws.cell(row=row, column=col).border = _border(GRAY_MID)
        
        ws.cell(row=row, column=col+1, value=t['moyenne'])
        ws.cell(row=row, column=col+1).number_format = '#,##0'
        ws.cell(row=row, column=col+1).font = Font(name="Arial", bold=True, size=9, color="1E293B")
        ws.cell(row=row, column=col+1).fill = _fill(GRAY_MID)
        ws.cell(row=row, column=col+1).alignment = _align("right")
        ws.cell(row=row, column=col+1).border = _border(GRAY_MID)
        
        ws.cell(row=row, column=col+2, value=t['prediction'])
        ws.cell(row=row, column=col+2).number_format = '#,##0'
        ws.cell(row=row, column=col+2).font = Font(name="Arial", bold=True, size=9, color=WHITE)
        ws.cell(row=row, column=col+2).fill = _fill(TEAL)
        ws.cell(row=row, column=col+2).alignment = _align("right")
        ws.cell(row=row, column=col+2).border = _border(TEAL)
        
        ws.cell(row=row, column=col+3, value=t['variation'] / 100)
        ws.cell(row=row, column=col+3).number_format = '+0.0%;-0.0%;0.0%'
        ws.cell(row=row, column=col+3).font = Font(name="Arial", bold=True, size=9, color=WHITE)
        ws.cell(row=row, column=col+3).fill = _fill(TEAL)
        ws.cell(row=row, column=col+3).alignment = _align("center")
        ws.cell(row=row, column=col+3).border = _border(TEAL)
        
        ws.row_dimensions[row].height = 18
    
    total_row = data_start_row + len(comptes)
    
    ws.cell(row=total_row, column=1, value="TOTAL GLOBAL")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws.cell(row=total_row, column=1).fill = _fill(NAVY)
    ws.cell(row=total_row, column=1).alignment = _align("center")
    ws.cell(row=total_row, column=1).border = _border(NAVY)
    
    col = 3
    for mois_data in comptes[0]['donnees_mensuelles'] if comptes else []:
        ws.cell(row=total_row, column=col, value="")
        ws.cell(row=total_row, column=col).fill = _fill(NAVY)
        ws.cell(row=total_row, column=col).border = _border(NAVY)
        
        ws.cell(row=total_row, column=col+1, value="")
        ws.cell(row=total_row, column=col+1).fill = _fill(NAVY)
        ws.cell(row=total_row, column=col+1).border = _border(NAVY)
        
        ws.cell(row=total_row, column=col+2, value="")
        ws.cell(row=total_row, column=col+2).fill = _fill(NAVY)
        ws.cell(row=total_row, column=col+2).border = _border(NAVY)
        
        ws.cell(row=total_row, column=col+3, value="")
        ws.cell(row=total_row, column=col+3).fill = _fill(NAVY)
        ws.cell(row=total_row, column=col+3).border = _border(NAVY)
        
        col += 4
    
    ws.cell(row=total_row, column=col, value=totaux_globaux['realise'])
    ws.cell(row=total_row, column=col).number_format = '#,##0'
    ws.cell(row=total_row, column=col).font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws.cell(row=total_row, column=col).fill = _fill(NAVY)
    ws.cell(row=total_row, column=col).alignment = _align("right")
    ws.cell(row=total_row, column=col).border = _border(NAVY)
    
    ws.cell(row=total_row, column=col+1, value=totaux_globaux['moyenne'])
    ws.cell(row=total_row, column=col+1).number_format = '#,##0'
    ws.cell(row=total_row, column=col+1).font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws.cell(row=total_row, column=col+1).fill = _fill(NAVY)
    ws.cell(row=total_row, column=col+1).alignment = _align("right")
    ws.cell(row=total_row, column=col+1).border = _border(NAVY)
    
    ws.cell(row=total_row, column=col+2, value=totaux_globaux['prediction'])
    ws.cell(row=total_row, column=col+2).number_format = '#,##0'
    ws.cell(row=total_row, column=col+2).font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws.cell(row=total_row, column=col+2).fill = _fill(NAVY)
    ws.cell(row=total_row, column=col+2).alignment = _align("right")
    ws.cell(row=total_row, column=col+2).border = _border(NAVY)
    
    ws.cell(row=total_row, column=col+3, value=totaux_globaux['variation'] / 100)
    ws.cell(row=total_row, column=col+3).number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(row=total_row, column=col+3).font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws.cell(row=total_row, column=col+3).fill = _fill(NAVY)
    ws.cell(row=total_row, column=col+3).alignment = _align("center")
    ws.cell(row=total_row, column=col+3).border = _border(NAVY)
    
    ws.row_dimensions[total_row].height = 24
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    for i in range(3, col + 4):
        ws.column_dimensions[get_column_letter(i)].width = 12
    
    ws.freeze_panes = f"C{sub_header_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(col+3)}{total_row-1}"


if __name__ == "__main__":
    print("Generation du fichier Excel...")
    
    output_path = os.path.join(project_root, "data")
    os.makedirs(output_path, exist_ok=True)
    
    print("\nClasse 6 (Charges):")
    try:
        data = generer_excel_predictions(classe=6, year_target=2027)
        path = os.path.join(output_path, "predictions_classe6_2027.xlsx")
        with open(path, "wb") as f:
            f.write(data)
        print(f"  Cree: {path}")
    except Exception as e:
        print(f"  Erreur: {e}")
    
    print("\nClasse 7 (Produits):")
    try:
        data = generer_excel_predictions(classe=7, year_target=2027)
        path = os.path.join(output_path, "predictions_classe7_2027.xlsx")
        with open(path, "wb") as f:
            f.write(data)
        print(f"  Cree: {path}")
    except Exception as e:
        print(f"  Erreur: {e}")
    
    print("\nTermine!")
